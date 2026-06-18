import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def export_fmea_to_excel(fmea_data, filename="fmea_output.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PFMEA"

    headers = ["Process Step", "Failure Mode", "Effect", "SEV",
               "Cause", "OCC", "Current Controls", "DET", "RPN",
               "Priority", "Recommended Actions"]

    header_fill = PatternFill("solid", fgColor="1D9E75")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    fills = {
        "HIGH":   PatternFill("solid", fgColor="F09595"),
        "MEDIUM": PatternFill("solid", fgColor="FAC775"),
        "LOW":    PatternFill("solid", fgColor="C0DD97")
    }

    for row_idx, row in enumerate(fmea_data["fmea_rows"], 2):
        values = [
            row["process_step"], row["failure_mode"], row["effect"],
            row["severity"], row["cause"], row["occurrence"],
            row["current_controls"], row["detection"], row["rpn"],
            row["priority"], row["recommended_actions"]
        ]
        priority = row["priority"]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fills.get(priority, fills["LOW"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["K"].width = 35
    ws.row_dimensions[1].height = 30

    wb.save(filename)
    print(f"FMEA exported to {filename}")
    return filename

if __name__ == "__main__":
    from fmea_rag import generate_fmea_with_context
    import json

    result = generate_fmea_with_context(
        process_name="CNC Milling — Aluminium Housing",
        process_steps=["Material loading", "Clamping",
                       "Rough milling", "Finish milling", "Inspection"],
        machine="DMG MORI DMU 50",
        material="Aluminium 6061-T6",
        quality_req=["Surface finish Ra 0.8", "Tolerance +/-0.05mm"],
        tolerances=["Bore diameter +/-0.02mm", "Flatness 0.05mm",
                    "Perpendicularity 0.03mm"],
        pdf_paths=["MTSK_Installation_guide_R02.pdf"]
    )

    export_fmea_to_excel(result, "fmea_output.xlsx")
    print("Open fmea_output.xlsx to see your FMEA!")